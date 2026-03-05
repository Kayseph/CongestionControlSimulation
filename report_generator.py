# report_generator.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ACCENT     = "ED7D31"
RED_HEX    = "C00000"
GREEN_HEX  = "375623"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"
DARK_GREY  = "595959"
MID_GREY   = "D9D9D9"

EVENT_COLORS = {
    "BUFFER_FULL":     ("C00000", "FFE0E0"),
    "AQM_DROP":        ("ED7D31", "FFF0E0"),
    "PHASE_CHANGE":    ("2E75B6", "E0EEFF"),
    "QUEUE_RECOVERED": ("375623", "E0FFE8"),
}

def _s(c=MID_GREY):       return Side(style="thin", color=c)
def _border(c=MID_GREY):  s = _s(c); return Border(left=s, right=s, top=s, bottom=s)
def _fill(h):             return PatternFill("solid", fgColor=h)
def _font(sz=10, bold=False, color="000000"):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=False)
def _col(ws, c, w): ws.column_dimensions[get_column_letter(c)].width = w

def _title(ws, row, c1, c2, text, bg=MID_BLUE, sz=11):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = _font(sz, True, WHITE)
    cell.fill = _fill(bg)
    cell.alignment = _center()
    ws.row_dimensions[row].height = 22

def _hdr(ws, row, headers, bg=MID_BLUE, h=30):
    for c, v in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = _font(10, True, WHITE)
        cell.fill = _fill(bg)
        cell.alignment = _center()
        cell.border = _border(DARK_BLUE)
    ws.row_dimensions[row].height = h

def _cell(ws, row, col, val, bg=WHITE, bold=False, color="000000",
          align="center", fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = _font(10, bold, color)
    cell.fill      = _fill(bg)
    cell.border    = _border()
    cell.alignment = _center() if align == "center" else _left()
    if fmt: cell.number_format = fmt
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Summary Dashboard
# ─────────────────────────────────────────────────────────────────────────────
def _summary(wb, runs):
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L2")
    b = ws["A1"]
    b.value = "Congestion Control Simulator — Full Simulation Report"
    b.font = _font(15, True, WHITE); b.fill = _fill(DARK_BLUE); b.alignment = _center()
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:L3")
    s = ws["A3"]
    s.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
               f"    |    Total Runs: {len(runs)}")
    s.font = _font(10, color=DARK_GREY); s.fill = _fill(LIGHT_GREY); s.alignment = _center()
    ws.row_dimensions[3].height = 16

    HDR = 5
    headers = ["Run #", "AQM", "Sender", "Flows",
               "Throughput\n(pkt/s)", "Loss Rate", "Avg Queue\n(pkts)",
               "Avg Delay\n(s)", "Fairness", "AQM Drops", "Buf Drops", "Status"]
    widths  = [7, 10, 9, 7, 15, 11, 14, 13, 11, 12, 12, 10]
    _hdr(ws, HDR, headers, h=36)
    for c, w in enumerate(widths, 1): _col(ws, c, w)

    for i, m in enumerate(runs):
        r  = HDR + 1 + i; ws.row_dimensions[r].height = 18
        bg = WHITE if i % 2 == 0 else LIGHT_GREY
        fair   = m["fairness"]
        status = "✓ Good" if fair >= 0.9 else ("⚠ Fair" if fair >= 0.7 else "✗ Poor")
        sc     = GREEN_HEX if fair >= 0.9 else (ACCENT if fair >= 0.7 else RED_HEX)
        vals = [i+1, m["aqm"], m["sender"], m["flows"],
                round(m["throughput"],2), round(m["loss_rate"],4),
                round(m["avg_queue"],2),  round(m["avg_delay"],4), round(fair,4),
                m.get("total_aqm_drops", 0), m.get("total_buf_drops", 0), status]
        for c, v in enumerate(vals, 1):
            cell = _cell(ws, r, c, v, bg, fmt="0.00%" if c==6 else None)
            if c == 12: cell.font = _font(10, True, sc)

    last = HDR + len(runs)
    KR   = last + 2
    _title(ws, KR, 1, 6, "📊  Aggregate Statistics", DARK_BLUE)
    _hdr(ws, KR+1, ["Metric","Min","Max","Average","Best Run","Notes"], h=20)
    metrics = [("Throughput (pkt/s)","E",False,"Higher = better"),
               ("Loss Rate","F",True,"Lower = better"),
               ("Avg Queue (pkts)","G",True,"Lower = better"),
               ("Avg Delay (s)","H",True,"Lower = better"),
               ("Fairness Index","I",False,"Closer to 1.0 = better")]
    d1, d2 = HDR+1, last
    for ki, (metric, col, lo, note) in enumerate(metrics):
        r  = KR+2+ki; ws.row_dimensions[r].height = 16
        bg = WHITE if ki % 2 == 0 else LIGHT_GREY
        rng  = f"{col}{d1}:{col}{d2}"
        best = (f"=INDEX(A{d1}:A{d2},MATCH(MIN({rng}),{rng},0))" if lo
                else f"=INDEX(A{d1}:A{d2},MATCH(MAX({rng}),{rng},0))")
        for c, v in enumerate([metric,f"=MIN({rng})",f"=MAX({rng})",f"=AVERAGE({rng})",best,note],1):
            _cell(ws, r, c, v, bg, fmt="0.00%" if metric=="Loss Rate" and c in(2,3,4) else None,
                  align="center" if c!=1 else "left")
    ws.freeze_panes = f"A{HDR+1}"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Congestion Story
# ─────────────────────────────────────────────────────────────────────────────
def _congestion_story(wb, runs):
    ws = wb.create_sheet("Congestion Story")
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 1, 14,
           "Congestion Build-up & Recovery — Annotated Step-by-Step", DARK_BLUE, sz=12)

    for c, w in enumerate([8,8,10,10,10,10,10,10,10,10,12,16,20,50], 1): _col(ws, c, w)

    row = 2
    for ri, m in enumerate(runs):
        ts = m["time_series"]; n = len(ts["time"])
        _title(ws, row, 1, 14,
               f"Run {ri+1}  |  AQM: {m['aqm']}  |  Sender: {m['sender']}  |  Flows: {m['flows']}",
               MID_BLUE); row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        lc = ws.cell(row=row, column=1,
            value="🔴 Buffer Overflow   🟠 AQM Drop   🔵 Phase Change   🟢 Queue Recovered  ← row colours")
        lc.font = _font(9, color=DARK_GREY); lc.fill = _fill(LIGHT_GREY); lc.alignment = _left()
        ws.row_dimensions[row].height = 13; row += 1

        _hdr(ws, row, ["Time\n(s)","Arrivals","Queue\n(pkts)","AQM\nAvg-Q","AQM\nDrop P",
                       "AQM\nDrops","Buf\nDrops","Throughput\n(pkt/s)","Loss\nRatio","Delay\n(s)",
                       "cwnd\n(Flow 0)","Phase\n(Flow 0)","Event Type","Event Detail"], h=36); row += 1

        ev_map = {}
        for t, etype, detail in m.get("events", []):
            ev_map.setdefault(round(t,2), []).append((etype, detail))

        sh     = m.get("sender_histories", [])
        cwnd_h = sh[0]["cwnd"]  if sh else []
        ph_h   = sh[0]["phase"] if sh else []

        for si in range(n):
            t_val = ts["time"][si]; evs = ev_map.get(round(t_val,2), [])
            bg = WHITE if si % 2 == 0 else LIGHT_GREY
            if evs: _, bg = EVENT_COLORS.get(evs[0][0], ("000000", bg))

            vals = [ts["time"][si], ts["arrivals"][si], ts["queue"][si],
                    ts["aqm_avg_queue"][si], ts["aqm_drop_prob"][si],
                    ts["aqm_drops"][si], ts["buffer_drops"][si],
                    round(ts["throughput"][si],2), round(ts["loss"][si],4),
                    round(ts["delay"][si],4),
                    cwnd_h[si] if si < len(cwnd_h) else "",
                    ph_h[si]   if si < len(ph_h)   else "",
                    " | ".join(e[0] for e in evs) if evs else "",
                    " | ".join(e[1] for e in evs) if evs else ""]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = _font(9); cell.fill = _fill(bg)
                cell.border = _border(); cell.alignment = _center() if c <= 12 else _left()
                if c == 9: cell.number_format = "0.00%"
                if c == 13 and evs:
                    tc, _ = EVENT_COLORS.get(evs[0][0], (DARK_GREY, ""))
                    cell.font = _font(9, True, tc)
            ws.row_dimensions[row].height = 13; row += 1
        row += 2


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — Narrative Summary
# Reads aggregates computed in network.py — never depends on event list length
# ─────────────────────────────────────────────────────────────────────────────
def _narrative(wb, runs):
    ws = wb.create_sheet("Narrative Summary")
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 1, 8,
           "Narrative Summary", DARK_BLUE, sz=12)
    _col(ws, 1, 12)
    for c in range(2, 9): _col(ws, c, 20)

    r = 3
    for ri, m in enumerate(runs):
        ts = m["time_series"]

        # ── Pull pre-computed aggregates directly from run dict ───────────────
        total_aqm = m.get("total_aqm_drops", 0)
        total_buf = m.get("total_buf_drops",  0)
        first_aqm = m.get("first_aqm_time")
        first_buf = m.get("first_buf_time")
        first_rec = m.get("first_rec_time")
        n_recov   = m.get("reno_recoveries",  0)
        first_r_t = m.get("first_recovery_t")
        n_ca      = m.get("reno_ca_count",    0)

        max_q    = max(ts["queue"]) if ts["queue"] else 0
        max_q_t  = ts["time"][ts["queue"].index(max_q)] if max_q else 0
        max_dp   = max(ts["aqm_drop_prob"]) if ts["aqm_drop_prob"] else 0
        avg_arr  = sum(ts["arrivals"]) / len(ts["arrivals"]) if ts["arrivals"] else 0

        lines = []

        # 1. Traffic build-up — always based on raw time-series
        lines.append(
            f"▶  TRAFFIC BUILD-UP:  {m['flows']} {m['sender']} flow(s) send aggressively "
            f"on a 10 Mbps link (avg {avg_arr:.0f} pkts/step, link capacity = 83 pkts/step).  "
            + (f"Queue peaked at {max_q} pkts at t={max_q_t}s — significant congestion developed.  "
               f"Total packets dropped (AQM + overflow): {total_aqm + total_buf}."
               if max_q >= 20
               else f"Queue peaked at {max_q} pkts — moderate congestion.  "
                    f"Total packets dropped: {total_aqm + total_buf}.")
        )

        # 2. AQM response — uses total_aqm/total_buf counts, NOT event list
        if total_aqm > 0:
            lines.append(
                f"▶  AQM RESPONSE ({m['aqm']}):  {total_aqm:,} packets dropped early "
                f"(first at t={first_aqm}s, peak drop-probability {max_dp:.1%}).  "
                + ("ARED dynamically increased max_p as queue pressure grew above max_th, "
                   "then relaxed it as flows backed off — adaptive feedback working correctly."
                   if m['aqm'] == 'ARED'
                   else "NLRED applied a quadratic probability ramp — gentle early signals "
                        "that steepen near max_th, giving flows proportional warning."
                   if m['aqm'] == 'NLRED'
                   else "RED used a linear drop-probability ramp between min_th=20 and max_th=60, "
                        "providing early congestion signals before the buffer reached capacity.")
            )
        elif total_buf > 0:
            lines.append(
                f"▶  AQM RESPONSE (DropTail):  {total_buf:,} packets hard-dropped at t={first_buf}s "
                f"when the buffer hit {BUFFER_SIZE} pkts.  "
                f"DropTail has no early-warning mechanism — senders only learn of congestion "
                f"after the buffer is already completely full, causing synchronised drops."
            )
        else:
            lines.append(
                f"▶  AQM ({m['aqm']}):  No drops were needed — traffic stayed within "
                f"link capacity throughout this run."
            )

        # 3. Sender response — uses pre-computed phase counters
        if m["sender"] == "RENO":
            if n_recov > 0:
                lines.append(
                    f"▶  SENDER RESPONSE (TCP RENO):  {n_recov} loss-triggered cwnd reductions.  "
                    f"Each time a drop signal arrived, RENO halved its congestion window "
                    f"(multiplicative decrease) and re-entered slow-start, then climbed back "
                    f"through congestion avoidance — the classic sawtooth pattern.  "
                    f"First reduction at t={first_r_t}s; {n_ca} congestion-avoidance "
                    f"transitions observed across the run."
                )
            else:
                lines.append(
                    f"▶  SENDER RESPONSE (TCP RENO):  cwnd grew continuously through slow-start "
                    f"and congestion avoidance without any loss-triggered reductions this run."
                )
        else:
            lines.append(
                f"▶  SENDER RESPONSE (BBR):  BBR paced packets at btlbw × rtprop — "
                f"loss-independent rate control.  It responded to rising RTT rather than "
                f"packet loss, keeping the queue from bloating while maintaining high throughput.  "
                f"BBR alternated between Probe BW (grow rate estimate) and Probe RTT phases."
            )

        # 4. Recovery
        if first_rec:
            lines.append(
                f"▶  RECOVERY:  Queue fell below 10 pkts at t={first_rec}s.  "
                f"The AQM early-drops and sender cwnd reductions successfully drained the "
                f"congested queue, restoring normal throughput."
            )
        elif total_aqm > 0 or total_buf > 0:
            lines.append(
                f"▶  RECOVERY:  Queue remained elevated throughout — the offered load "
                f"continuously exceeded capacity.  AQM and senders managed the congestion "
                f"but could not fully drain it given {m['flows']} competing flows."
            )
        else:
            lines.append(f"▶  RECOVERY:  No congestion — no recovery needed.")

        # 5. Final outcome
        lines.append(
            f"▶  FINAL OUTCOME:  Throughput {m['throughput']:.2f} pkt/s  |  "
            f"Loss rate {m['loss_rate']:.4f}  |  Avg queue {m['avg_queue']:.1f} pkts  |  "
            f"Avg delay {m['avg_delay']:.4f}s  |  Fairness {m['fairness']:.4f}."
        )

        _title(ws, r, 1, 8,
               f"Run {ri+1}  —  {m['aqm']} / {m['sender']} / {m['flows']} Flows  "
               f"|  {m.get('timestamp','')}",
               MID_BLUE); r += 1

        BKG = {"▶  TRAFFIC": LIGHT_GREY, "▶  AQM":     LIGHT_BLUE,
               "▶  SENDER":  "F0F0FF",   "▶  RECOVERY":"E8FFE8",
               "▶  FINAL":   LIGHT_GREY}
        for line in lines:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            cell = ws.cell(row=r, column=1, value=line)
            cell.font      = _font(10)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border    = _border()
            cell.fill      = _fill(next((v for k,v in BKG.items() if line.startswith(k)), LIGHT_GREY))
            ws.row_dimensions[r].height = 50; r += 1
        r += 1


BUFFER_SIZE = 100   # matches network.py — used in narrative text only


def generate_report(runs_meta: list, output_dir: str = ".") -> str:
    if not runs_meta:
        raise ValueError("No simulation runs to report on.")
    wb = Workbook()
    _summary(wb, runs_meta)
    _congestion_story(wb, runs_meta)
    _narrative(wb, runs_meta)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"SimulationReport_{ts}.xlsx")
    wb.save(filepath)
    return filepath
